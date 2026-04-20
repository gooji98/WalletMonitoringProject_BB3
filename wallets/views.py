import csv
from decimal import Decimal
from io import BytesIO

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font

from .models import Wallet, BalanceSnapshot, TransactionSnapshot
from .sync import sync_all_active_wallets, sync_wallet


User = get_user_model()

ALLOWED_PAGE_SIZES = {"10", "20", "100"}


def get_per_page(request, key, default=10):
    raw = request.GET.get(key, str(default))
    return int(raw) if raw in ALLOWED_PAGE_SIZES else default


def get_eth_usd_rate():
    return Decimal(str(getattr(settings, "ETH_USD_RATE", 3000)))


def get_usd_class(value):
    if value is None:
        return "text-muted"
    if value >= Decimal("100000"):
        return "text-success fw-semibold"
    if value >= Decimal("10000"):
        return "text-primary fw-semibold"
    if value >= Decimal("1000"):
        return "text-warning-emphasis fw-semibold"
    return "text-body"


def build_wallet_rows(wallets_queryset):
    eth_usd_rate = get_eth_usd_rate()
    now = timezone.now()
    rows = []

    for wallet in wallets_queryset:
        latest_balance = wallet.balance_snapshots.order_by("-fetched_at").first()
        balance_eth = latest_balance.balance_eth if latest_balance else None
        balance_usd = (balance_eth * eth_usd_rate) if balance_eth is not None else None
        last_balance_update = latest_balance.fetched_at if latest_balance else None

        is_stale = False
        if last_balance_update:
            is_stale = (now - last_balance_update).total_seconds() > 3600

        rows.append({
            "wallet": wallet,
            "latest_balance": latest_balance,
            "balance_eth": balance_eth,
            "balance_usd": balance_usd,
            "balance_usd_class": get_usd_class(balance_usd),
            "last_balance_update": last_balance_update,
            "is_stale": is_stale,
        })

    return rows


def get_dashboard_wallets_queryset(request):
    wallet_q = request.GET.get("wallet_q", "").strip()
    wallet_status = request.GET.get("wallet_status", "").strip()
    wallet_network = request.GET.get("wallet_network", "").strip()
    wallet_admin = request.GET.get("wallet_admin", "").strip()

    wallets = Wallet.objects.select_related("assigned_admin").order_by("-created_at")

    if wallet_q:
        wallets = wallets.filter(
            Q(label__icontains=wallet_q)
            | Q(address__icontains=wallet_q)
            | Q(network__icontains=wallet_q)
            | Q(assigned_admin__username__icontains=wallet_q)
            | Q(assigned_admin__email__icontains=wallet_q)
        )

    if wallet_status == "active":
        wallets = wallets.filter(is_active=True)
    elif wallet_status == "inactive":
        wallets = wallets.filter(is_active=False)

    if wallet_network:
        wallets = wallets.filter(network=wallet_network)

    if wallet_admin:
        wallets = wallets.filter(assigned_admin_id=wallet_admin)

    return wallets


def get_dashboard_wallets_page(request):
    wallets_queryset = get_dashboard_wallets_queryset(request)
    per_page = get_per_page(request, "wallet_per_page", 10)
    paginator = Paginator(wallets_queryset, per_page)
    page_number = request.GET.get("wallet_page")
    page_obj = paginator.get_page(page_number)

    wallet_rows = build_wallet_rows(page_obj.object_list)

    return wallets_queryset, page_obj, wallet_rows, per_page


def get_dashboard_transactions_queryset(request):
    tx_q = request.GET.get("dtx_q", "").strip()
    tx_status = request.GET.get("dtx_status", "").strip()

    transactions = (
        TransactionSnapshot.objects
        .select_related("wallet", "wallet__assigned_admin")
        .order_by("-tx_timestamp")
    )

    if tx_q:
        transactions = transactions.filter(
            Q(tx_hash__icontains=tx_q)
            | Q(from_address__icontains=tx_q)
            | Q(to_address__icontains=tx_q)
            | Q(wallet__label__icontains=tx_q)
            | Q(wallet__address__icontains=tx_q)
        )

    if tx_status == "success":
        transactions = transactions.filter(is_error=False)
    elif tx_status == "error":
        transactions = transactions.filter(is_error=True)

    return transactions


def get_dashboard_transactions_page(request):
    transactions_queryset = get_dashboard_transactions_queryset(request)
    per_page = get_per_page(request, "dtx_per_page", 10)
    paginator = Paginator(transactions_queryset, per_page)
    page_number = request.GET.get("dtx_page")
    page_obj = paginator.get_page(page_number)

    return transactions_queryset, page_obj, per_page


def get_wallet_snapshots_queryset(wallet, request):
    snapshot_q = request.GET.get("snapshot_q", "").strip()

    snapshots = wallet.balance_snapshots.order_by("-fetched_at")

    if snapshot_q:
        snapshots = snapshots.filter(
            Q(balance_wei__icontains=snapshot_q)
            | Q(balance_eth__icontains=snapshot_q)
            | Q(fetched_at__icontains=snapshot_q)
        )

    return snapshots


def get_wallet_snapshots_page(wallet, request):
    snapshots_queryset = get_wallet_snapshots_queryset(wallet, request)
    per_page = get_per_page(request, "snapshot_per_page", 10)
    paginator = Paginator(snapshots_queryset, per_page)
    page_number = request.GET.get("snapshot_page")
    page_obj = paginator.get_page(page_number)

    return snapshots_queryset, page_obj, per_page


def get_wallet_transactions_queryset(wallet, request):
    tx_q = request.GET.get("tx_q", "").strip()
    tx_status = request.GET.get("tx_status", "").strip()

    transactions = wallet.transaction_snapshots.order_by("-tx_timestamp")

    if tx_q:
        transactions = transactions.filter(
            Q(tx_hash__icontains=tx_q)
            | Q(from_address__icontains=tx_q)
            | Q(to_address__icontains=tx_q)
        )

    if tx_status == "success":
        transactions = transactions.filter(is_error=False)
    elif tx_status == "error":
        transactions = transactions.filter(is_error=True)

    return transactions


def get_wallet_transactions_page(wallet, request):
    transactions_queryset = get_wallet_transactions_queryset(wallet, request)
    per_page = get_per_page(request, "tx_per_page", 10)
    paginator = Paginator(transactions_queryset, per_page)
    page_number = request.GET.get("tx_page")
    page_obj = paginator.get_page(page_number)

    return transactions_queryset, page_obj, per_page


def dashboard(request):
    wallets_queryset, wallets_page, wallet_rows, wallet_per_page = get_dashboard_wallets_page(request)
    dashboard_tx_queryset, dashboard_tx_page, dtx_per_page = get_dashboard_transactions_page(request)

    networks = Wallet.objects.order_by("network").values_list("network", flat=True).distinct()
    admins = User.objects.filter(managed_wallets__isnull=False).distinct().order_by("username")

    all_wallet_rows = build_wallet_rows(wallets_queryset[:50])
    pie_labels = [row["wallet"].label or row["wallet"].address[:10] for row in all_wallet_rows if row["balance_eth"] is not None]
    pie_values = [float(row["balance_eth"]) for row in all_wallet_rows if row["balance_eth"] is not None]

    context = {
        "total_wallets": Wallet.objects.count(),
        "active_wallets": Wallet.objects.filter(is_active=True).count(),
        "total_balance_snapshots": BalanceSnapshot.objects.count(),
        "total_transactions": TransactionSnapshot.objects.count(),

        "wallet_rows": wallet_rows,
        "wallets_page": wallets_page,
        "wallet_per_page": wallet_per_page,
        "wallet_q": request.GET.get("wallet_q", "").strip(),
        "wallet_status": request.GET.get("wallet_status", "").strip(),
        "wallet_network": request.GET.get("wallet_network", "").strip(),
        "wallet_admin": request.GET.get("wallet_admin", "").strip(),
        "wallet_networks": networks,
        "wallet_admins": admins,

        "dashboard_recent_transactions": dashboard_tx_page,
        "dtx_per_page": dtx_per_page,
        "dtx_q": request.GET.get("dtx_q", "").strip(),
        "dtx_status": request.GET.get("dtx_status", "").strip(),

        "dashboard_pie_labels": pie_labels,
        "dashboard_pie_values": pie_values,
    }

    return render(request, "wallets/dashboard.html", context)


def dashboard_wallets_partial(request):
    _, wallets_page, wallet_rows, wallet_per_page = get_dashboard_wallets_page(request)

    networks = Wallet.objects.order_by("network").values_list("network", flat=True).distinct()
    admins = User.objects.filter(managed_wallets__isnull=False).distinct().order_by("username")

    context = {
        "wallet_rows": wallet_rows,
        "wallets_page": wallets_page,
        "wallet_per_page": wallet_per_page,
        "wallet_q": request.GET.get("wallet_q", "").strip(),
        "wallet_status": request.GET.get("wallet_status", "").strip(),
        "wallet_network": request.GET.get("wallet_network", "").strip(),
        "wallet_admin": request.GET.get("wallet_admin", "").strip(),
        "wallet_networks": networks,
        "wallet_admins": admins,
    }

    return render(request, "wallets/partials/dashboard_wallets_table.html", context)


def dashboard_transactions_partial(request):
    _, dashboard_tx_page, dtx_per_page = get_dashboard_transactions_page(request)

    context = {
        "dashboard_recent_transactions": dashboard_tx_page,
        "dtx_per_page": dtx_per_page,
        "dtx_q": request.GET.get("dtx_q", "").strip(),
        "dtx_status": request.GET.get("dtx_status", "").strip(),
    }

    return render(request, "wallets/partials/dashboard_transactions_table.html", context)


def sync_now(request):
    sync_all_active_wallets()
    messages.success(request, "Wallet sync completed successfully.")
    return redirect("dashboard")


def wallet_detail(request, wallet_id):
    wallet = get_object_or_404(Wallet.objects.select_related("assigned_admin"), id=wallet_id)

    latest_balance = wallet.balance_snapshots.order_by("-fetched_at").first()
    _, snapshots_page, snapshot_per_page = get_wallet_snapshots_page(wallet, request)
    _, transactions_page, tx_per_page = get_wallet_transactions_page(wallet, request)

    balance_history_for_chart = list(wallet.balance_snapshots.order_by("-fetched_at")[:10])[::-1]

    chart_labels = [
        snapshot.fetched_at.strftime("%Y-%m-%d %H:%M")
        for snapshot in balance_history_for_chart
    ]
    chart_values = [
        float(snapshot.balance_eth)
        for snapshot in balance_history_for_chart
    ]

    eth_usd_rate = get_eth_usd_rate()
    latest_balance_usd = (latest_balance.balance_eth * eth_usd_rate) if latest_balance else None

    total_tx_count = wallet.transaction_snapshots.count()
    success_tx_count = wallet.transaction_snapshots.filter(is_error=False).count()
    error_tx_count = wallet.transaction_snapshots.filter(is_error=True).count()

    wallet_pie_labels = ["Success", "Error"]
    wallet_pie_values = [success_tx_count, error_tx_count]

    context = {
        "wallet": wallet,
        "latest_balance": latest_balance,
        "latest_balance_usd": latest_balance_usd,

        "recent_snapshots": snapshots_page,
        "snapshot_per_page": snapshot_per_page,
        "snapshot_q": request.GET.get("snapshot_q", "").strip(),

        "recent_transactions": transactions_page,
        "tx_per_page": tx_per_page,
        "tx_q": request.GET.get("tx_q", "").strip(),
        "tx_status": request.GET.get("tx_status", "").strip(),

        "chart_labels": chart_labels,
        "chart_values": chart_values,
        "wallet_pie_labels": wallet_pie_labels,
        "wallet_pie_values": wallet_pie_values,
        "total_tx_count": total_tx_count,
    }

    return render(request, "wallets/wallet_detail.html", context)


def wallet_snapshots_partial(request, wallet_id):
    wallet = get_object_or_404(Wallet, id=wallet_id)
    _, snapshots_page, snapshot_per_page = get_wallet_snapshots_page(wallet, request)

    context = {
        "wallet": wallet,
        "recent_snapshots": snapshots_page,
        "snapshot_per_page": snapshot_per_page,
        "snapshot_q": request.GET.get("snapshot_q", "").strip(),
    }

    return render(request, "wallets/partials/wallet_snapshots_table.html", context)


def wallet_transactions_partial(request, wallet_id):
    wallet = get_object_or_404(Wallet, id=wallet_id)
    _, transactions_page, tx_per_page = get_wallet_transactions_page(wallet, request)

    context = {
        "wallet": wallet,
        "recent_transactions": transactions_page,
        "tx_per_page": tx_per_page,
        "tx_q": request.GET.get("tx_q", "").strip(),
        "tx_status": request.GET.get("tx_status", "").strip(),
    }

    return render(request, "wallets/partials/wallet_transactions_table.html", context)


def export_wallet_transactions_csv(request, wallet_id):
    wallet = get_object_or_404(Wallet, id=wallet_id)
    transactions = wallet.transaction_snapshots.order_by("-tx_timestamp")

    safe_label = (wallet.label or f"wallet_{wallet.id}").replace(" ", "_")
    filename = f"{safe_label}_transactions.csv"

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow([
        "wallet_id",
        "wallet_label",
        "wallet_address",
        "tx_hash",
        "from_address",
        "to_address",
        "value_wei",
        "value_eth",
        "block_number",
        "tx_timestamp",
        "is_error",
        "fetched_at",
    ])

    for tx in transactions:
        writer.writerow([
            wallet.id,
            wallet.label or "",
            wallet.address,
            tx.tx_hash,
            tx.from_address,
            tx.to_address or "",
            tx.value_wei,
            tx.value_eth,
            tx.block_number,
            tx.tx_timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            tx.is_error,
            tx.fetched_at.strftime("%Y-%m-%d %H:%M:%S"),
        ])

    return response


def export_wallet_snapshots_csv(request, wallet_id):
    wallet = get_object_or_404(Wallet, id=wallet_id)
    snapshots = wallet.balance_snapshots.order_by("-fetched_at")

    safe_label = (wallet.label or f"wallet_{wallet.id}").replace(" ", "_")
    filename = f"{safe_label}_balance_snapshots.csv"

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow([
        "wallet_id",
        "wallet_label",
        "wallet_address",
        "balance_wei",
        "balance_eth",
        "fetched_at",
    ])

    for snapshot in snapshots:
        writer.writerow([
            wallet.id,
            wallet.label or "",
            wallet.address,
            snapshot.balance_wei,
            snapshot.balance_eth,
            snapshot.fetched_at.strftime("%Y-%m-%d %H:%M:%S"),
        ])

    return response


def export_dashboard_summary_xlsx(request):
    wallets_queryset = get_dashboard_wallets_queryset(request)
    dashboard_transactions_queryset = get_dashboard_transactions_queryset(request)

    wb = Workbook()

    ws_wallets = wb.active
    ws_wallets.title = "Wallets"

    wallet_headers = [
        "wallet_id",
        "label",
        "address",
        "network",
        "assigned_admin",
        "is_active",
        "created_at",
        "last_synced_at",
        "latest_balance_eth",
        "latest_balance_usd",
        "latest_balance_fetched_at",
    ]
    ws_wallets.append(wallet_headers)

    eth_usd_rate = get_eth_usd_rate()

    for wallet in wallets_queryset.select_related("assigned_admin"):
        latest_balance = wallet.balance_snapshots.order_by("-fetched_at").first()
        latest_balance_eth = latest_balance.balance_eth if latest_balance else None
        latest_balance_usd = (latest_balance_eth * eth_usd_rate) if latest_balance_eth is not None else None

        ws_wallets.append([
            wallet.id,
            wallet.label or "",
            wallet.address,
            wallet.network,
            wallet.assigned_admin.username if wallet.assigned_admin else "",
            wallet.is_active,
            wallet.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            wallet.last_synced_at.strftime("%Y-%m-%d %H:%M:%S") if wallet.last_synced_at else "",
            str(latest_balance_eth) if latest_balance_eth is not None else "",
            str(latest_balance_usd) if latest_balance_usd is not None else "",
            latest_balance.fetched_at.strftime("%Y-%m-%d %H:%M:%S") if latest_balance else "",
        ])

    ws_transactions = wb.create_sheet(title="Recent Transactions")
    tx_headers = [
        "wallet_id",
        "wallet_label",
        "wallet_address",
        "network",
        "assigned_admin",
        "tx_hash",
        "from_address",
        "to_address",
        "value_wei",
        "value_eth",
        "block_number",
        "tx_timestamp",
        "is_error",
        "fetched_at",
    ]
    ws_transactions.append(tx_headers)

    for tx in dashboard_transactions_queryset.select_related("wallet", "wallet__assigned_admin"):
        ws_transactions.append([
            tx.wallet.id,
            tx.wallet.label or "",
            tx.wallet.address,
            tx.wallet.network,
            tx.wallet.assigned_admin.username if tx.wallet.assigned_admin else "",
            tx.tx_hash,
            tx.from_address,
            tx.to_address or "",
            tx.value_wei,
            str(tx.value_eth),
            tx.block_number,
            tx.tx_timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            tx.is_error,
            tx.fetched_at.strftime("%Y-%m-%d %H:%M:%S"),
        ])

    for ws in [ws_wallets, ws_transactions]:
        for cell in ws[1]:
            cell.font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="dashboard_summary.xlsx"'
    return response


def sync_single_wallet(request, wallet_id):
    wallet = get_object_or_404(Wallet, id=wallet_id)
    sync_wallet(wallet)
    messages.success(request, f"Wallet '{wallet.label or wallet.address}' synced successfully.")
    return redirect("wallet_detail", wallet_id=wallet.id)